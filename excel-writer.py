from openpyxl import Workbook

from Test import DOC, TestCase ,TC_Stub

from Test import TC_Stub

wb = Workbook()

ctf_file = open("test.tcf", "r")
doc = DOC()
doc.extract_doc_atrributes(ctf_file.read())
ctf_file = open("test.tcf", "r")
TestCase(doc).extract_test_case(ctf_file.read())
doc.print_information()

# grab the active worksheet
ws = wb.active

ws['B2'] = 'XLStoTCF_ISIT'
ws['C2'] = doc.version

ws.merge_cells('B4:F4')
ws['B4'] = 'Test Sequence Attributes'

ws['B5'] = 'Sequence Name'
ws['C5'] = doc.sequenceName

ws['B6'] = 'Sequence Documentation'
ws['C6'] = doc.sequenceDocumentation

ws['B7'] = 'Developer Name'
ws['C7'] = doc.developerName

ws['B8'] = 'Tester Name'
ws['C8'] = doc.testerName

ws['B9'] = 'Creation Date'
ws['C9'] = doc.creationDate

ws['B10'] = 'Procedure'
ws['C10'] = doc.procedure

ws['B11'] = 'Procedure Number'
ws['C11'] = doc.procedureNumber

ws['B12'] = 'Source File'
ws['C12'] = doc.sourceFile

ws['B13'] = 'Language'
ws['C13'] = doc.language

ws['B14'] = 'Coverage Mode'
ws['C14'] = doc.coverageMode

ws['B15'] = 'Additional Files or Procedures'
ws['C15'] = doc.additionalFiles

ws['B55'] = 'Test Case Data Sets'

ws['B17'] = 'Test Sequence Code Inserts'

ws['B22'] = 'Pre Include Code'

ws['B25'] = 'Post Include Code'


ws['B31'] = 'File Based Test Case Cleanup Code'

ws['B34'] = 'Globals Environment'

ws['B34'] = 'Stubs Environment'



current = 56


for idx, variable in enumerate(doc.test_cases[0].input_variables):
    row = 56 + idx * 6
    pos = 'B' + str(row)
    ws[str(pos)] = 'Input Variable ' + str(idx + 1)

    pos = 'C' + str(row + 1)
    ws[str(pos)] = 'Name '

    pos = 'C' + str(row + 2)
    ws[str(pos)] = 'Type '

    pos = 'C' + str(row + 3)
    ws[str(pos)] = 'Usage '

    pos = 'C' + str(row + 4)
    ws[str(pos)] = 'Value '

    pos = 'C' + str(row + 5)
    ws[str(pos)] = 'Value Retained ? '

    current = row

for idx, variable in enumerate(doc.test_cases[0].output_variables):
    row = current + (idx+1) * 7
    pos = 'B' + str(row)
    ws[str(pos)] = 'Output Variable' + str(idx + 1)

    pos = 'C' + str(row + 1)
    ws[str(pos)] = 'Name '

    pos = 'C' + str(row + 2)
    ws[str(pos)] = 'Type '

    pos = 'C' + str(row + 3)
    ws[str(pos)] = 'Usage '

    pos = 'C' + str(row + 4)
    ws[str(pos)] = 'Value '

    pos = 'C' + str(row + 5)
    ws[str(pos)] = 'Comparison '

    pos = 'C' + str(row + 6)
    ws[str(pos)] = 'TBrun Analysis'
nbr=len(doc.test_cases)
print(nbr)

nbr= len(doc.stubs) // len(doc.test_cases)

C='B'
l=309
print(nbr)
for idx in range(nbr):
    ws['B'+str(l)] = 'Test Case Stub '+str(idx+1)
    ws['C'+str(l+1)] ='Procedure'
    ws['C' + str(l + 2)] = 'Name'
    ws['B' + str(l + 3)] = 'TC Hit Count'
    ws['C' + str(l + 4)] = 'Value'
    ws['B' + str(l + 5)] = 'TC Hit Order'
    ws['C' + str(l + 6)] = 'Value'
    ws['B'+ str(l + 7)] = 'Input Parameter 1'
    ws['C' + str(l + 8)] = 'Name'
    ws['C' + str(l + 9)] = 'Type'
    ws['C' + str(l + 10)] = 'Value'
    ws['C' + str(l + 11)] = 'Comparaison'
    ws['B' + str(l + 12)] = 'Return Value'
    ws['C' + str(l + 13)]= 'Name'
    ws['C' + str(l + 14)] = 'Type'
    ws['C' + str(l + 15)] = 'Value'
    l=l+18

i=0
j=38


ws['B'+str(37)]='Stubs Environment'
ws['C'+str(37)]='Procedure'
ws['D'+str(37)]='Method'
ws['E'+str(37)]='Method'
ws['F'+str(37)]='Default Return Value'
ws['G'+str(37)]='File'

for st in doc.stub_file:

    i=i+1
    ws['C'+str(j)]=st.Procedure+'('+str(st.overloading)+')'
    ws['D'+str(j)]=st.Method
    ws['B'+str(j)]='Stub '+str(i)
    j=j+1


col = 'D'
col2 = 'A'
j=0


for test_case in doc.test_cases:


      row=57
      j = j + 1
      ws[col + str(44)]='Test Case'+str(j)
      ws['B45'] = 'Test Case Attributes'
      ws[col + str(46)] = test_case.Name
      ws[col + str(47)] = test_case.Documentation
      j=0
      k=0
      for input_variable in test_case.input_variables:
          j=j+1
          ws[col + str(row)] =  input_variable.Name
          ws[col + str(row + 1)] = input_variable.Type
          ws[col + str(row + 2)] = 'input global'
          ws[col + str(row + 3)] = input_variable.Value

          row = row + 6

      row = row+1
      print(row)

      for output_variable in test_case.output_variables:
          k=k+1
          ws[col + str(row)] = output_variable.Name
          ws[col + str(row+1)] = output_variable.Type
          if (output_variable.Usage=='H'):
              ws[col + str(row+2)] = 'output global'
          if (output_variable.Usage=='O'):
               ws[col + str(row + 2)] = 'output Parameter'

          ws[col + str(row+3)] = output_variable.Value
          ws[col + str(row+4)] = '!='
          ws[col + str(row +5)] = 'Compare + Write'

          row = row + 7
      col=chr(ord(col)+1)

















wb.save('sample.xlsx')

